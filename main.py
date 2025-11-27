import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

class ProcesadorExcel:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de Archivos Excel")
        self.root.geometry("700x400")
        self.carpeta_seleccionada = ""
        
        # Columnas requeridas en el orden especificado
        self.columnas_salida = [
            "FOLIO", "NOMBRE", "RUT", "FECHA DE FIRMA", "DIRECCION", "COMUNA", "NACIONALIDAD",
            "ESTADO CIVIL", "FECHA DE NACIMIENTO", "CIUDAD", "PLANTA", "CARGO",
            "FUNCION", "LUGAR DE PRESTACION", "SUELDO BASE", "CODIGO GRATIFICACION",
            "COLACION", "MOVILIZACION", "FECHA DE INICIO", "FECHA 1º VENCIMIENTO",
            "FECHA DE 2º VENCIMIENTO", "INDEFINIDO", "FONASA O ISAPRE", "JORNADA LABORAL",
            "CORREO", "N°TELEFONO", "FECHA DE PAGO DE REMUNERACION"
        ]

        # Mapeo de columnas originales a columnas de salida (por nombre)
        self.mapeo_columnas = {
            "Nombre ": "NOMBRE", "Nombre": "NOMBRE", "NOMBRE": "NOMBRE", "Nombre Completo": "NOMBRE",
            "Rut": "RUT", "RUT": "RUT",
            "Inicio Contrato": "FECHA DE FIRMA", "Fecha de Firma": "FECHA DE FIRMA",
            "FECHA DE FIRMA": "FECHA DE FIRMA", "Fecha Firma": "FECHA DE FIRMA",
            "Dirección": "DIRECCION", "Direccion": "DIRECCION", "DIRECCION": "DIRECCION",
            "Comuna": "COMUNA", "COMUNA": "COMUNA",
            "Nacionalidad": "NACIONALIDAD", "NACIONALIDAD": "NACIONALIDAD",
            "EstCiv": "ESTADO CIVIL", "Estado Civil": "ESTADO CIVIL", "ESTADO CIVIL": "ESTADO CIVIL",
            "Fecha Nac.": "FECHA DE NACIMIENTO", "Fecha de Nacimiento": "FECHA DE NACIMIENTO",
            "FECHA DE NACIMIENTO": "FECHA DE NACIMIENTO", "Fecha Nacimiento": "FECHA DE NACIMIENTO",
            "Ciudad": "CIUDAD", "CIUDAD": "CIUDAD",
            "Planta": "PLANTA", "PLANTA": "PLANTA", "Sucursal": "PLANTA", "Descripción.1": "PLANTA",
            "Cargo": "CARGO", "CARGO": "CARGO",
            "Descripción.3": "FUNCION", "Funcion": "FUNCION", "Función": "FUNCION", "FUNCION": "FUNCION",
            "Unidad Organizativa": "LUGAR DE PRESTACION", "Descripción.2": "LUGAR DE PRESTACION",
            "Lugar de Prestacion": "LUGAR DE PRESTACION", "Lugar de Prestación": "LUGAR DE PRESTACION",
            "LUGAR DE PRESTACION": "LUGAR DE PRESTACION",
            "Nombre de LPS": "LUGAR DE PRESTACION",
            "Importe": "SUELDO BASE", "Sueldo Base": "SUELDO BASE", "SUELDO BASE": "SUELDO BASE", "Sueldo": "SUELDO BASE",
            "Importe.1": "COLACION", "Colacion": "COLACION", "Colación": "COLACION", "COLACION": "COLACION",
            "Importe.2": "MOVILIZACION", "Movilizacion": "MOVILIZACION",
            "Movilización": "MOVILIZACION", "MOVILIZACION": "MOVILIZACION",
            "Fecha de Inicio": "FECHA DE INICIO", "FECHA DE INICIO": "FECHA DE INICIO", "Fecha Inicio": "FECHA DE INICIO",
            "Rec. Antigüedad Emp.": "FECHA 1º VENCIMIENTO", "Fecha 1º Vencimiento": "FECHA 1º VENCIMIENTO",
            "FECHA 1º VENCIMIENTO": "FECHA 1º VENCIMIENTO", "Fecha 1 Vencimiento": "FECHA 1º VENCIMIENTO",
            "Fin Contrato": "FECHA DE 2º VENCIMIENTO", "Fin Estimado": "FECHA DE 2º VENCIMIENTO",
            "Fecha 2º Vencimiento": "FECHA DE 2º VENCIMIENTO",
            "FECHA 2º VENCIMIENTO": "FECHA DE 2º VENCIMIENTO", "Fecha 2 Vencimiento": "FECHA DE 2º VENCIMIENTO",
            "Sistema Salud": "FONASA O ISAPRE", "Fonasa o Isapre": "FONASA O ISAPRE",
            "FONASA O ISAPRE": "FONASA O ISAPRE",
            "Prevision": "FONASA O ISAPRE", "Previsión": "FONASA O ISAPRE", "Salud": "FONASA O ISAPRE",
            "Jornada": "JORNADA LABORAL", "Descripción.4": "JORNADA LABORAL",
            "Jornada Laboral": "JORNADA LABORAL", "JORNADA LABORAL": "JORNADA LABORAL",
            "Mail": "CORREO", "Mail Personal": "CORREO", "Correo": "CORREO",
            "CORREO": "CORREO", "Email": "CORREO",
            "Nº teléfono": "N°TELEFONO", "Telefono": "N°TELEFONO", "Teléfono": "N°TELEFONO",
            "TELEFONO": "N°TELEFONO", "N° Telefono": "N°TELEFONO",
            "N°TELEFONO": "N°TELEFONO", "Numero Telefono": "N°TELEFONO", "Celular": "N°TELEFONO"
        }

        self.crear_interfaz()
    
    def crear_interfaz(self):
        frame_botones = tk.Frame(self.root, pady=10)
        frame_botones.pack(fill=tk.X)
        
        btn_carpeta = tk.Button(frame_botones, text="Elegir Carpeta", 
                                command=self.elegir_carpeta, width=15, height=2)
        btn_carpeta.pack(side=tk.LEFT, padx=10)
        
        btn_procesar = tk.Button(frame_botones, text="Procesar", 
                                 command=self.procesar_archivos, width=15, height=2)
        btn_procesar.pack(side=tk.LEFT, padx=10)
        
        btn_cerrar = tk.Button(frame_botones, text="Cerrar", 
                               command=self.root.quit, width=15, height=2)
        btn_cerrar.pack(side=tk.LEFT, padx=10)
        
        frame_log = tk.Frame(self.root)
        frame_log.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tk.Label(frame_log, text="Log de Operaciones:", anchor=tk.W).pack(fill=tk.X)
        
        self.log_text = scrolledtext.ScrolledText(frame_log, wrap=tk.WORD, 
                                                   height=15, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
    
    def escribir_log(self, mensaje):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{mensaje}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update()
    
    def elegir_carpeta(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta con archivos Excel")
        if carpeta:
            self.carpeta_seleccionada = carpeta
            self.escribir_log(f"Carpeta seleccionada: {carpeta}")
        else:
            self.escribir_log("No se seleccionó ninguna carpeta.")
    
    def formatear_columna_fecha(self, serie):
        """
        Convierte una serie a formato dd-mm-yyyy, deja vacío si no es fecha válida.
        """
        fechas = pd.to_datetime(serie, errors='coerce', dayfirst=True)
        resultado = fechas.dt.strftime("%d-%m-%Y")
        return resultado.where(~fechas.isna(), "")
    
    def mapear_datos(self, df_original):
        num_filas = len(df_original)
        datos_mapeados = {}

        # Inicializar todas las columnas vacías
        for col in self.columnas_salida:
            datos_mapeados[col] = [""] * num_filas

        hoy_str = datetime.now().strftime("%d-%m-%Y")

        # FOLIO → 1, 2, 3, ...
        datos_mapeados["FOLIO"] = list(range(1, num_filas + 1))

        # FECHA DE FIRMA = HOY
        datos_mapeados["FECHA DE FIRMA"] = [hoy_str] * num_filas

        # Forzados por posición (índice 0-based)
        overrides = [
            ("LUGAR DE PRESTACION", 82, "columna CE"),
            ("NACIONALIDAD", 12, "columna M"),
            ("COLACION", 79, "columna CB"),
            ("MOVILIZACION", 76, "columna BY"),
            ("CODIGO GRATIFICACION", 71, "columna BT"),
            ("PLANTA", 21, "columna V"),
            ("INDEFINIDO", 32, "columna AG")
        ]

        for col_salida, idx, desc in overrides:
            try:
                datos_mapeados[col_salida] = df_original.iloc[:, idx].tolist()
                self.escribir_log(f"      ✓ Forzado: {desc} → {col_salida}")
            except Exception as e:
                self.escribir_log(f"      ✗ No se pudo asignar {desc} a {col_salida}: {str(e)}")

        # Mapeo normal por nombre (solo si la columna destino sigue vacía)
        for col_original in df_original.columns:
            if col_original in self.mapeo_columnas:
                col_salida = self.mapeo_columnas[col_original]
                if datos_mapeados[col_salida] == [""] * num_filas:
                    datos_mapeados[col_salida] = df_original[col_original].tolist()
                    self.escribir_log(f"      ✓ Mapeado: '{col_original}' → '{col_salida}'")
        
        df_salida = pd.DataFrame(datos_mapeados)

        # NOMBRE compuesto
        if all(c in df_original.columns for c in ["Nombre de pila", "Ap.Paterno", "Apellido de soltera"]):
            nombre_compuesto = (
                df_original["Nombre de pila"].fillna("").astype(str).str.strip() + " " +
                df_original["Ap.Paterno"].fillna("").astype(str).str.strip() + " " +
                df_original["Apellido de soltera"].fillna("").astype(str).str.strip()
            )
            df_salida["NOMBRE"] = nombre_compuesto.str.replace(r"\s+", " ", regex=True).str.strip()
            self.escribir_log("      ✓ NOMBRE compuesto desde Nombre de pila + Ap.Paterno + Apellido de soltera")

        # 🆕 ESTADO CIVIL: agregar " o/a" al final (si no está vacío)
        if "ESTADO CIVIL" in df_salida.columns:
            df_salida["ESTADO CIVIL"] = df_salida["ESTADO CIVIL"].apply(
                lambda x: f"{str(x).strip()}o/a" if pd.notna(x) and str(x).strip() != "" else ""
            )

        # FECHA DE INICIO copia FECHA DE FIRMA si venía completamente vacía
        if df_salida["FECHA DE INICIO"].equals(pd.Series([""] * num_filas)):
            df_salida["FECHA DE INICIO"] = df_salida["FECHA DE FIRMA"]
            self.escribir_log("      ✓ FECHA DE INICIO copiada desde FECHA DE FIRMA")

        # Fecha de pago vacía
        df_salida["FECHA DE PAGO DE REMUNERACION"] = ""

        # Formatear columnas de fecha
        columnas_fecha = [
            "FECHA DE FIRMA",
            "FECHA DE NACIMIENTO",
            "FECHA DE INICIO",
            "FECHA 1º VENCIMIENTO",
            "FECHA DE 2º VENCIMIENTO",
            "FECHA DE PAGO DE REMUNERACION",
        ]
        for col in columnas_fecha:
            if col in df_salida.columns:
                df_salida[col] = self.formatear_columna_fecha(df_salida[col])

        # Asegurar orden de columnas
        df_salida = df_salida[self.columnas_salida]

        return df_salida
    
    def aplicar_formato(self, archivo_salida):
        try:
            wb = load_workbook(archivo_salida)
            ws = wb.active
            
            fill_celeste = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            font_negrita = Font(bold=True)

            for cell in ws[1]:
                cell.fill = fill_celeste
                cell.font = font_negrita
            
            wb.save(archivo_salida)
        except Exception as e:
            self.escribir_log(f"  ⚠ Advertencia: No se pudo aplicar formato: {str(e)}")
    
    def procesar_archivos(self):
        if not self.carpeta_seleccionada:
            messagebox.showwarning("Advertencia", "Por favor, seleccione una carpeta primero.")
            return
        
        self.escribir_log("\n" + "="*70)
        self.escribir_log("INICIANDO PROCESAMIENTO")
        self.escribir_log("="*70)
        
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        carpeta_salida = os.path.join(self.carpeta_seleccionada, f"Procesados_{fecha_actual}")

        try:
            os.makedirs(carpeta_salida, exist_ok=True)
            self.escribir_log(f"Carpeta de salida: {carpeta_salida}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear carpeta: {str(e)}")
            return
        
        try:
            todos_archivos = os.listdir(self.carpeta_seleccionada)
            archivos_excel = [
                f for f in todos_archivos
                if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~$')
            ]
            
            if not archivos_excel:
                messagebox.showinfo("Información", "No hay archivos Excel para procesar.")
                self.escribir_log("No se encontraron archivos Excel para procesar.")
                return
            
            self.escribir_log(f"Archivos encontrados: {len(archivos_excel)}")

        except Exception as e:
            self.escribir_log(f"✗ Error al listar archivos: {str(e)}")
            return
        
        archivos_procesados = 0
        archivos_error = 0
        
        for archivo in archivos_excel:
            try:
                ruta_entrada = os.path.join(self.carpeta_seleccionada, archivo)
                nombre_base = os.path.splitext(archivo)[0]
                ruta_salida = os.path.join(carpeta_salida, f"{nombre_base}_procesado.xlsx")
                
                self.escribir_log(f"\n📄 Procesando: {archivo}")
                
                df = pd.read_excel(ruta_entrada)
                self.escribir_log(f"  ✓ {len(df)} filas, {len(df.columns)} columnas")
                
                df_procesado = self.mapear_datos(df)
                
                df_procesado.to_excel(ruta_salida, index=False, engine='openpyxl')
                self.aplicar_formato(ruta_salida)
                
                archivos_procesados += 1

            except Exception as e:
                archivos_error += 1
                self.escribir_log(f"  ✗ ERROR procesando '{archivo}': {str(e)}")
        
        self.escribir_log("\n=== RESUMEN ===")
        self.escribir_log(f"✓ Procesados: {archivos_procesados}")
        self.escribir_log(f"✗ Errores: {archivos_error}")
        self.escribir_log(f"📁 Guardados en: {carpeta_salida}")
        
        messagebox.showinfo(
            "Completado",
            f"Procesamiento finalizado.\n\nExitosos: {archivos_procesados}\nErrores: {archivos_error}"
        )

def main():
    root = tk.Tk()
    app = ProcesadorExcel(root)
    root.mainloop()

if __name__ == "__main__":
    main()
